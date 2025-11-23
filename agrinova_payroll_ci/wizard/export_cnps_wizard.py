# -*- coding: utf-8 -*-
from odoo import models, fields, api
import base64
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None


class ExportCNPSWizard(models.TransientModel):
    _name = 'export.cnps.wizard'
    _description = 'Export Déclaration CNPS'
    
    payslip_ids = fields.Many2many(
        'hr.payslip',
        string='Bulletins de paie',
        required=True
    )
    
    date_from = fields.Date(
        string='Date de début',
        required=True
    )
    
    date_to = fields.Date(
        string='Date de fin',
        required=True
    )
    
    export_file = fields.Binary(
        string='Fichier Excel',
        readonly=True
    )
    
    export_filename = fields.Char(
        string='Nom du fichier',
        readonly=True
    )
    
    def action_export_cnps(self):
        """Génère le fichier Excel d'export CNPS"""
        if not Workbook:
            raise UserError("La bibliothèque openpyxl n'est pas installée. Installez-la avec: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Déclaration CNPS"
        
        # En-têtes
        headers = [
            'Numéro CNPS',
            'Nom',
            'Prénom',
            'Salaire Brut',
            'Base CNPS',
            'CNPS Employé (6.3%)',
            'CNPS Employeur (16.55%)',
            'Total CNPS'
        ]
        
        # Style en-têtes
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        row_num = 2
        total_brut = 0
        total_cnps_emp = 0
        total_cnps_empl = 0
        
        for payslip in self.payslip_ids:
            ws.cell(row=row_num, column=1, value=payslip.contract_id.cnps_number or '')
            # Séparation Nom / Prénom basique
            name_parts = (payslip.employee_id.name or '').partition(' ')
            ws.cell(row=row_num, column=2, value=name_parts[0]) # Nom (premier mot)
            ws.cell(row=row_num, column=3, value=name_parts[2]) # Prénom (reste)
            ws.cell(row=row_num, column=4, value=payslip.contract_id.wage)
            ws.cell(row=row_num, column=5, value=payslip.contract_id.cnps_base)
            ws.cell(row=row_num, column=6, value=payslip.cnps_employee)
            ws.cell(row=row_num, column=7, value=payslip.cnps_employer)
            ws.cell(row=row_num, column=8, value=payslip.cnps_employee + payslip.cnps_employer)
            
            total_brut += payslip.contract_id.wage
            total_cnps_emp += payslip.cnps_employee
            total_cnps_empl += payslip.cnps_employer
            
            row_num += 1
        
        # Totaux
        ws.cell(row=row_num, column=3, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row_num, column=4, value=total_brut).font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=total_cnps_emp).font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=total_cnps_empl).font = Font(bold=True)
        ws.cell(row=row_num, column=8, value=total_cnps_emp + total_cnps_empl).font = Font(bold=True)
        
        # Ajuster largeur colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Sauvegarder
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.export_file = base64.b64encode(output.read())
        self.export_filename = f'CNPS_Export_{self.date_from}_{self.date_to}.xlsx'
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.cnps.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
