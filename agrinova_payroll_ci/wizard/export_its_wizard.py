# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None


class ExportITSWizard(models.TransientModel):
    _name = 'export.its.wizard'
    _description = 'Export Déclaration ITS'
    
    payslip_ids = fields.Many2many(
        'hr.payslip',
        string='Bulletins de paie',
        required=True
    )
    
    date_from = fields.Date(
        string='Date de début',
        required=True
    )
    
    date_to = fields.Date(
        string='Date de fin',
        required=True
    )
    
    export_file = fields.Binary(
        string='Fichier Excel',
        readonly=True
    )
    
    export_filename = fields.Char(
        string='Nom du fichier',
        readonly=True
    )
    
    def action_export_its(self):
        """Génère le fichier Excel d'export ITS"""
        if not Workbook:
            raise UserError("La bibliothèque openpyxl n'est pas installée.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Déclaration ITS"
        
        # En-têtes
        headers = [
            'Matricule',
            'Nom Complet',
            'Salaire Brut',
            'CNPS Employé',
            'Net Imposable',
            'ITS',
            'Net à Payer'
        ]
        
        # Style
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        row_num = 2
        total_brut = 0
        total_cnps = 0
        total_its = 0
        total_net = 0
        
        for payslip in self.payslip_ids:
            brut = payslip.contract_id.wage
            net_pay_line = payslip.line_ids.filtered(lambda l: l.code == 'NET')
            net_pay = net_pay_line.total if net_pay_line else 0
            
            ws.cell(row=row_num, column=1, value=payslip.employee_id.identification_id or '')
            ws.cell(row=row_num, column=2, value=payslip.employee_id.name)
            ws.cell(row=row_num, column=3, value=brut)
            ws.cell(row=row_num, column=4, value=payslip.cnps_employee)
            ws.cell(row=row_num, column=5, value=payslip.net_imposable)
            ws.cell(row=row_num, column=6, value=payslip.its_amount)
            ws.cell(row=row_num, column=7, value=net_pay)
            
            total_brut += brut
            total_cnps += payslip.cnps_employee
            total_its += payslip.its_amount
            total_net += net_pay
            
            row_num += 1
        
        # Totaux
        ws.cell(row=row_num, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row_num, column=3, value=total_brut).font = Font(bold=True)
        ws.cell(row=row_num, column=4, value=total_cnps).font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=total_its).font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=total_net).font = Font(bold=True)
        
        # Ajuster colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        # Sauvegarder
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.export_file = base64.b64encode(output.read())
        self.export_filename = f'ITS_Export_{self.date_from}_{self.date_to}.xlsx'
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.its.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
